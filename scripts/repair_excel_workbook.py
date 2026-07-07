from __future__ import annotations

import argparse
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from backend.utils.serialization.excel_exporter import ExcelExporter


def _sheet_headers(ws) -> List[Any]:
    if ws.max_row < 1:
        return []
    return [cell.value for cell in ws[1]]


def _replace_sheet(wb, title: str):
    if title in wb.sheetnames:
        del wb[title]
    return wb.create_sheet(title)


def _record_from_row(headers: List[str], row: tuple[Any, ...], row_number: int, exporter: ExcelExporter) -> Dict[str, Any]:
    record = {
        header: "" if idx >= len(row) or row[idx] is None else row[idx]
        for idx, header in enumerate(headers)
        if header
    }
    for header in exporter.PIPELINE_HEADERS:
        record.setdefault(header, "")
    record["_row_number"] = row_number
    record["_event_type_raw"] = record.get("event_type", "")
    return record


def _blank_non_applicable_metrics(clean: Dict[str, Any], exporter: ExcelExporter) -> None:
    event_type = str(clean.get("event_type", "")).upper()
    groups = []
    if "ENGINE" not in event_type:
        groups.extend(["engine_score", "engine_depth", "engine_ms"])
    if "VISION" not in event_type:
        groups.extend(["detections_count", "avg_confidence", "min_confidence", "yolo_latency_ms", "yolo_fps"])
    if "ROBOT" not in event_type:
        groups.extend(["robot_ms"])
    for field in groups:
        clean[field] = ""


def _clean_record(record: Dict[str, Any], exporter: ExcelExporter) -> Dict[str, Any]:
    clean = dict(record)
    clean["event_type"] = exporter._canonical_event_type(clean.get("event_type", "UNKNOWN"))
    raw = exporter._load_json(clean.get("raw_payload"))
    raw = raw if isinstance(raw, dict) else {}

    for field in ("camera_status", "engine_score", "engine_depth", "ai_move", "fen_after", "ucci_position"):
        suggested, has_suggestion = exporter._raw_field_suggestion(clean, raw, field)
        if has_suggestion:
            clean[field] = "" if suggested is None else suggested

    _blank_non_applicable_metrics(clean, exporter)
    exporter._apply_audit(clean)
    return clean


def repair_workbook(path: Path) -> Path:
    if not path.exists():
        raise FileNotFoundError(path)

    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup = path.with_name(f"{path.stem}.before_excel_fix_{timestamp}{path.suffix}")
    shutil.copy2(path, backup)

    exporter = ExcelExporter(filename=str(path), subscribe=False)
    wb = load_workbook(path)
    if exporter.PIPELINE_SHEET not in wb.sheetnames:
        raise ValueError(f"missing sheet: {exporter.PIPELINE_SHEET}")

    ws = wb[exporter.PIPELINE_SHEET]
    headers = _sheet_headers(ws)
    if headers not in (exporter.PIPELINE_BASE_HEADERS, exporter.PIPELINE_HEADERS):
        raise ValueError(f"unexpected Pipeline_Log headers: {headers}")

    for header in exporter.PIPELINE_HEADERS[len(headers):]:
        ws.cell(row=1, column=ws.max_column + 1, value=header)
    headers = _sheet_headers(ws)
    header_index = {header: idx + 1 for idx, header in enumerate(headers)}

    clean_rows: List[Dict[str, Any]] = []
    quality_rows: List[Dict[str, Any]] = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        record = _record_from_row(headers, row, row_number, exporter)
        exporter._apply_audit(record)
        ws.cell(row=row_number, column=header_index["audit_status"], value=record.get("audit_status", ""))
        ws.cell(row=row_number, column=header_index["audit_notes"], value=record.get("audit_notes", ""))
        quality_rows.extend(exporter._quality_issues(record))
        clean_rows.append(_clean_record(record, exporter))

    clean_ws = _replace_sheet(wb, "Clean_Pipeline_Log")
    clean_ws.append(exporter.PIPELINE_HEADERS)
    for record in clean_rows:
        clean_ws.append([record.get(header, "") for header in exporter.PIPELINE_HEADERS])

    quality_ws = _replace_sheet(wb, "Data Quality")
    quality_ws.append(exporter.DATA_QUALITY_HEADERS)
    for row in quality_rows:
        quality_ws.append([row.get(header, "") for header in exporter.DATA_QUALITY_HEADERS])

    for sheet in (ws, clean_ws, quality_ws):
        exporter._style_sheet(sheet)

    wb.save(path)
    wb.close()
    print(f"backup={backup}")
    print(f"clean_rows={len(clean_rows)}")
    print(f"data_quality_rows={len(quality_rows)}")
    return backup


def main() -> int:
    parser = argparse.ArgumentParser(description="Repair Excel audit and clean views.")
    parser.add_argument("path", nargs="?", default="chess_robot_experiment.xlsx")
    args = parser.parse_args()
    repair_workbook(Path(args.path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

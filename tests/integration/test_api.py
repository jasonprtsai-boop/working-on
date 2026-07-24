import os
import uuid
import unittest
from io import BytesIO

from openpyxl import load_workbook


class TestApiRoutes(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        os.environ["FAKE_VISION"] = "1"
        from backend.utils import config

        config.FAKE_VISION = True
        from backend.main import create_app

        cls.app, _socketio = create_app()
        cls.client = cls.app.test_client()
        login = cls.client.post("/api/login", json={"username": "admin", "password": config.ADMIN_PASSWORD})
        token = (login.get_json() or {}).get("token")
        cls.auth_headers = {"Authorization": f"Bearer {token}"}

    def test_state_endpoint_exposes_legacy_aliases(self):
        resp = self.client.get("/api/state", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json() or {}
        self.assertIn("game", payload)
        self.assertIn("board", payload)
        self.assertIn("state", payload)
        self.assertIn("history", payload)
        self.assertIn("fen", payload)

    def test_player_state_is_public_and_redacted(self):
        started = self.client.post(
            "/api/runtime/session/start",
            json={"participant_id": "P-SECRET"},
            headers=self.auth_headers,
        )
        self.assertEqual(started.status_code, 200)
        try:
            resp = self.app.test_client().get("/api/player/state")
            self.assertEqual(resp.status_code, 200)
            payload = resp.get_json() or {}
            self.assertIn("board", payload)
            self.assertIn("sync", payload)
            self.assertIn("ui", payload)
            self.assertNotIn("game", payload)
            self.assertNotIn("participant_id", payload.get("ui", {}))
            self.assertNotIn("session_id", payload.get("ui", {}))
            self.assertNotIn("record_path", payload.get("ui", {}))
            self.assertNotIn("position", payload.get("robot", {}))
        finally:
            self.client.post("/api/runtime/session/end", json={}, headers=self.auth_headers)

    def test_sensitive_read_endpoints_require_auth(self):
        client = self.app.test_client()
        for path in (
            "/api/state",
            "/api/health",
            "/api/vision/status",
            "/api/video_status",
            "/api/engine/status",
            "/api/estop/status",
            "/api/vision/cameras",
            "/api/vision/source/status",
            "/api/vision/calibration",
            "/api/runtime/status",
            "/api/robot/calibration",
            "/api/runtime/metrics",
            "/api/runtime/control",
            "/api/assets/status",
            "/api/video_feed",
            "/api/vision/stream",
            "/api/vision/snapshot",
            "/api/snapshot",
            "/api/export/excel",
            "/api/replay/sessions",
            "/api/replay/steps",
            "/api/replay/step/0",
            "/api/replay/export",
        ):
            with self.subTest(path=path):
                resp = client.get(path)
                self.assertEqual(resp.status_code, 401)
                self.assertEqual((resp.get_json() or {}).get("code"), "unauthorized")

    def test_engine_status_endpoint_is_available(self):
        resp = self.client.get("/api/engine/status", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json() or {}
        self.assertIn("status", payload)
        self.assertIn("report", payload)

    def test_vision_status_exposes_single_protected_yolo_model(self):
        resp = self.client.get("/api/vision/status", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json() or {}
        model_report = ((payload.get("models") or {}).get("vision_models") or {})
        model = model_report.get("model") or {}

        self.assertEqual(model.get("extension"), ".onnx")
        self.assertTrue(model.get("protected"))
        self.assertTrue((model_report.get("dataset_mapping") or {}).get("names_match_nc"))

    def test_runtime_status_exposes_workers_queues_and_event_bus(self):
        resp = self.client.get("/api/runtime/status", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json() or {}
        self.assertIn("workers", payload)
        self.assertIn("queues", payload)
        self.assertIn("queue", payload)
        self.assertIn("event_bus", payload)
        self.assertIn("async_runtime", payload)
        self.assertIn("persistence", payload)
        self.assertIn("runtime", payload)
        self.assertEqual(payload.get("queue"), payload.get("queues"))
        self.assertIn("global_subscribers", payload.get("event_bus", {}))
        self.assertIn("dropped_events", payload.get("persistence", {}))

    def test_runtime_metrics_exposes_compact_apm_counters(self):
        resp = self.client.get("/api/runtime/metrics", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json() or {}
        self.assertIn("timestamp", payload)
        self.assertIn("status_counts", payload.get("workers", {}))
        self.assertIn("event_bus", payload)
        self.assertIn("async_runtime", payload)
        self.assertIn("persistence", payload)
        self.assertIn("queues", payload)

    def test_runtime_control_updates_depth_safe_mode_and_session(self):
        from backend.runtime.workers.engine_worker import engine_worker

        resp = self.client.post("/api/runtime/engine-depth", json={"depth": 20}, headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        payload = resp.get_json() or {}
        self.assertEqual(payload.get("engine_depth"), 20)
        self.assertEqual(engine_worker.depth_on_change, 20)
        self.assertEqual(engine_worker.depth_on_idle, 20)

        mode = self.client.post("/api/runtime/ai-mode", json={"mode": "companionship"}, headers=self.auth_headers)
        self.assertEqual(mode.status_code, 200)
        mode_payload = mode.get_json() or {}
        self.assertEqual(mode_payload.get("ai_mode"), "companionship")
        self.assertEqual(mode_payload.get("ai_mode_label"), "陪伴模式")
        self.assertEqual(mode_payload.get("engine_depth"), 6)
        self.assertEqual(engine_worker.depth_on_change, 6)

        safe = self.client.post("/api/runtime/safe-mode", json={"enabled": False}, headers=self.auth_headers)
        self.assertEqual(safe.status_code, 200)
        self.assertFalse((safe.get_json() or {}).get("safe_mode"))

        started = self.client.post(
            "/api/runtime/session/start",
            json={"participant_id": "P-TEST"},
            headers=self.auth_headers,
        )
        self.assertEqual(started.status_code, 200)
        session = (started.get_json() or {}).get("session") or {}
        self.assertTrue(session.get("active"))
        self.assertEqual(session.get("participant_id"), "P-TEST")

        ended = self.client.post("/api/runtime/session/end", json={}, headers=self.auth_headers)
        self.assertEqual(ended.status_code, 200)
        self.assertFalse(((ended.get_json() or {}).get("session") or {}).get("active"))

    def test_excel_export_returns_research_workbook(self):
        denied = self.app.test_client().get("/api/export/excel")
        self.assertEqual(denied.status_code, 401)

        resp = self.client.get("/api/export/excel", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        self.assertIn("spreadsheet", resp.mimetype)

        data = resp.data
        resp.close()
        wb = load_workbook(BytesIO(data), read_only=True)
        try:
            for sheet_name in ("Overview", "Pipeline_Log", "Data Quality", "Vision FEN Log", "Vision Detections", "UCCI Trace"):
                self.assertIn(sheet_name, wb.sheetnames)
            pipeline_headers = [cell.value for cell in wb["Pipeline_Log"][1]]
            self.assertIn("audit_status", pipeline_headers)
            self.assertIn("audit_notes", pipeline_headers)
        finally:
            wb.close()

    def test_csv_export_returns_runtime_event_rows(self):
        resp = self.client.get("/api/export/csv", headers=self.auth_headers)
        self.assertEqual(resp.status_code, 200)
        self.assertEqual(resp.mimetype, "text/csv")
        text = resp.get_data(as_text=True)
        resp.close()
        self.assertIn("sequence_id,session_id,trace_id,type,timestamp,payload_json", text)

    def test_replay_routes_expose_sessions_steps_snapshot_and_export(self):
        from backend.events.store.event_store import event_store

        session_id = f"replay-test-{uuid.uuid4().hex}"
        start_fen = "rnbakabnr/9/1c5c1/p1p1p1p1p/9/9/P1P1P1P1P/1C5C1/9/RNBAKABNR w - - 0 1"
        event_store.append({
            "session_id": session_id,
            "trace_id": "trace-replay-test",
            "type": "STATE_UPDATED",
            "payload": {
                "game": {
                    "fen": start_fen,
                    "move_history": ["b2b5"],
                    "current_turn": "b",
                }
            },
            "timestamp": 1234.0,
        })

        sessions = self.client.get("/api/replay/sessions", headers=self.auth_headers)
        self.assertEqual(sessions.status_code, 200)
        session_payload = sessions.get_json() or {}
        self.assertTrue(any(item.get("session_id") == session_id for item in session_payload.get("sessions", [])))

        steps = self.client.get(f"/api/replay/steps?session={session_id}", headers=self.auth_headers)
        self.assertEqual(steps.status_code, 200)
        step_payload = steps.get_json() or {}
        self.assertEqual(step_payload.get("total"), 1)
        self.assertEqual(step_payload["steps"][0]["move"], "b2b5")
        self.assertEqual(step_payload["steps"][0]["trace_id"], "trace-replay-test")

        snapshot = self.client.get(f"/api/replay/step/0?session={session_id}", headers=self.auth_headers)
        self.assertEqual(snapshot.status_code, 200)
        snapshot_payload = snapshot.get_json() or {}
        self.assertEqual((snapshot_payload.get("_replay") or {}).get("session_id"), session_id)
        self.assertEqual((snapshot_payload.get("board") or {}).get("fen"), start_fen)

        exported = self.client.get(f"/api/replay/export?session={session_id}", headers=self.auth_headers)
        self.assertEqual(exported.status_code, 200)
        self.assertEqual(exported.mimetype, "application/json")
        export_payload = exported.get_json() or {}
        self.assertEqual(export_payload.get("count"), 1)
        self.assertEqual(export_payload.get("session_id"), session_id)


if __name__ == "__main__":
    unittest.main()

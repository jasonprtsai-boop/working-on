import os
import subprocess
import sys
import tempfile
import unittest
import uuid
from pathlib import Path

from openpyxl import load_workbook

from backend.utils.serialization.excel_exporter import ExcelExporter, sanitize_excel_cell


class FakeState:
    def to_dict(self):
        return {
            "session_id": "session-a",
            "board": {"fen": "fen-from-state"},
            "game": {"phase": "RUNNING"},
            "sync": {
                "timeline": {
                    "vision": {"duration": 42},
                    "engine": {"duration": 120},
                    "robot": {"duration": 330},
                }
            },
            "engine": {"score": 0.25, "depth": 8},
            "robot": {"status": "idle"},
            "vision": {"fps": 28, "status": "OK"},
            "system": {"status": "RUNNING"},
        }


class TestExcelExporter(unittest.TestCase):
    def test_sanitize_excel_cell_escapes_formula_prefixes(self):
        for value in ("=2+2", "+SUM(A1:A2)", "-10+cmd", "@cmd", "  =trimmed"):
            self.assertEqual(sanitize_excel_cell(value), "'" + value)
        self.assertEqual(sanitize_excel_cell("safe"), "safe")
        self.assertEqual(sanitize_excel_cell(42), 42)

    def test_export_session_builds_research_workbook(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = os.path.join(tmpdir, "source.xlsx")
            out_path = os.path.join(tmpdir, "report.xlsx")
            exporter = ExcelExporter(filename=log_path, subscribe=False)

            exporter.log_event(
                FakeState(),
                {
                    "type": "VISION_MOVE_DETECTED",
                    "source": "test",
                    "trace_id": "trace-1",
                    "data": {
                        "session_id": "session-a",
                        "fen": "fen-from-yolo",
                        "ucci_position": "position fen fen-from-yolo",
                        "board_state": {"0,0": "R"},
                        "latency_ms": 1250,
                        "detections": [
                            {
                                "class_name": "red_rook",
                                "confidence": 0.42,
                                "bbox": [1, 2, 3, 4],
                                "cell": {"key": "0,0"},
                            }
                        ],
                        "detections_count": 1,
                        "avg_confidence": 0.42,
                        "min_confidence": 0.42,
                    },
                },
            )
            exporter.log_event(
                FakeState(),
                {
                    "type": "DIAGNOSTICS.UPDATED",
                    "source": "diagnostics",
                    "trace_id": "trace-diag",
                    "data": {
                        "session_id": "session-a",
                        "vision": {"camera_status": "READY"},
                    },
                },
            )
            exporter.log_event(
                FakeState(),
                {
                    "type": "ENGINE_ANALYSIS_COMPLETED",
                    "source": "engine",
                    "data": {
                        "session_id": "session-a",
                        "engine": {
                            "bestmove": "a0a1",
                            "score": 0.5,
                            "depth": 10,
                        },
                        "engine_ms": 90,
                    },
                },
            )
            exporter.log_event(
                FakeState(),
                {
                    "type": "SYSTEM_ERROR",
                    "source": "test",
                    "data": {"session_id": "session-a", "error": "camera offline"},
                },
            )
            exporter.log_event(
                FakeState(),
                {
                    "type": "VISION_BENCHMARK_RESULT",
                    "source": "vision_benchmark",
                    "data": {
                        "session_id": "session-a",
                        "mode": "full_yolo",
                        "frame_id": 1,
                        "status": "ok",
                        "detections_count": 2,
                        "small_object_count": 1,
                        "small_object_rate": 0.5,
                        "fps": 12.0,
                        "inference_latency_ms": 70.0,
                        "end_to_end_latency_ms": 83.0,
                        "fen": "9/9/9/9/9/9/9/9/9/9 w - - 0 1",
                        "fen_valid": True,
                        "stable_update": False,
                        "detections_json": "[]",
                        "board_state_json": "{}",
                        "map_50": "N/A",
                        "recall": "N/A",
                        "metric_note": "requires_annotations",
                        "requires_annotations": True,
                    },
                },
            )

            exporter.export_session("session-a", out_path)
            wb = load_workbook(out_path, data_only=True)
            try:
                expected = {
                    "Overview",
                    "Pipeline_Log",
                    "Data Quality",
                    "Vision FEN Log",
                    "Vision Detections",
                    "Vision Mode Comparison",
                    "UCCI Trace",
                    "Errors & Warnings",
                    "Raw Payload",
                }
                self.assertTrue(expected.issubset(set(wb.sheetnames)))
                self.assertEqual(wb["Vision FEN Log"].max_row, 2)
                self.assertEqual(wb["Vision Detections"]["C2"].value, "red_rook")

                pipeline_headers = [cell.value for cell in wb["Pipeline_Log"][1]]
                self.assertIn("audit_status", pipeline_headers)
                self.assertIn("audit_notes", pipeline_headers)
                event_id_col = pipeline_headers.index("event_id") + 1
                engine_score_col = pipeline_headers.index("engine_score") + 1
                event_type_col = pipeline_headers.index("event_type") + 1
                camera_status_col = pipeline_headers.index("camera_status") + 1

                first_event_id = wb["Pipeline_Log"].cell(row=2, column=event_id_col).value
                uuid.UUID(str(first_event_id))

                diag_row = next(
                    row for row in range(2, wb["Pipeline_Log"].max_row + 1)
                    if wb["Pipeline_Log"].cell(row=row, column=event_type_col).value == "DIAGNOSTICS_UPDATED"
                )
                self.assertEqual(wb["Pipeline_Log"].cell(row=diag_row, column=camera_status_col).value, "READY")
                self.assertIsNone(wb["Pipeline_Log"].cell(row=diag_row, column=engine_score_col).value)

                engine_row = next(
                    row for row in range(2, wb["Pipeline_Log"].max_row + 1)
                    if wb["Pipeline_Log"].cell(row=row, column=event_type_col).value == "ENGINE_ANALYSIS_COMPLETED"
                )
                self.assertEqual(wb["Pipeline_Log"].cell(row=engine_row, column=engine_score_col).value, 0.5)

                warning_reasons = [
                    wb["Errors & Warnings"][f"B{row}"].value
                    for row in range(2, wb["Errors & Warnings"].max_row + 1)
                ]
                self.assertIn("Low YOLO confidence", warning_reasons)
                self.assertIn("High YOLO latency", warning_reasons)
                self.assertIn("camera offline", warning_reasons)

                quality_fields = [
                    wb["Data Quality"][f"D{row}"].value
                    for row in range(2, wb["Data Quality"].max_row + 1)
                ]
                self.assertIn("event_id", quality_fields)
                self.assertIn("event_type", quality_fields)

                comparison = wb["Vision Mode Comparison"]
                comparison_headers = [cell.value for cell in comparison[1]]
                self.assertIn("avg_fps", comparison_headers)
                avg_fps_col = comparison_headers.index("avg_fps") + 1
                map_col = comparison_headers.index("map_50") + 1
                recall_col = comparison_headers.index("recall") + 1
                self.assertEqual(comparison["A2"].value, "summary")
                self.assertEqual(comparison["B2"].value, "full_yolo")
                self.assertEqual(comparison.cell(row=2, column=avg_fps_col).value, 12.0)
                self.assertEqual(comparison.cell(row=2, column=map_col).value, "N/A")
                self.assertEqual(comparison.cell(row=2, column=recall_col).value, "N/A")
                self.assertEqual(comparison["A3"].value, "detail")

                self.assertEqual(wb["Pipeline_Log"].freeze_panes, "A2")
                self.assertEqual(wb["Vision Mode Comparison"].sheet_properties.tabColor.rgb, "007C3AED")
                self.assertIn("A1:B1", [str(item) for item in wb["Overview"].merged_cells.ranges])
                self.assertEqual(wb["Overview"]["A1"].fill.fgColor.rgb, "00EFF6FF")
                self.assertEqual(comparison["A2"].fill.fgColor.rgb, "00DBEAFE")
            finally:
                wb.close()

    def test_empty_session_still_exports_all_report_sheets(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = os.path.join(tmpdir, "source.xlsx")
            out_path = os.path.join(tmpdir, "empty.xlsx")
            exporter = ExcelExporter(filename=log_path, subscribe=False)

            exporter.export_session("missing", out_path)
            wb = load_workbook(out_path, data_only=True)
            try:
                for sheet_name in ExcelExporter.REPORT_SHEETS:
                    self.assertIn(sheet_name, wb.sheetnames)
            finally:
                wb.close()

    def test_import_has_no_default_workbook_side_effect(self):
        project_root = Path(__file__).resolve().parents[2]
        with tempfile.TemporaryDirectory() as tmpdir:
            env = os.environ.copy()
            env["PYTHONPATH"] = str(project_root)
            script = (
                "import os; "
                f"os.chdir({tmpdir!r}); "
                "import backend.utils.serialization.excel_exporter; "
                "print(os.path.exists('chess_robot_experiment.xlsx'))"
            )
            result = subprocess.run(
                [sys.executable, "-c", script],
                check=True,
                text=True,
                capture_output=True,
                env=env,
            )
            self.assertEqual(result.stdout.strip(), "False")

    def test_formula_like_values_are_written_as_text(self):
        with tempfile.TemporaryDirectory() as tmpdir:
            log_path = os.path.join(tmpdir, "source.xlsx")
            out_path = os.path.join(tmpdir, "report.xlsx")
            exporter = ExcelExporter(filename=log_path, subscribe=False)

            exporter.log_event(
                FakeState(),
                {
                    "type": "GAME_PLAYER_MOVE",
                    "source": "@socket",
                    "trace_id": "+trace",
                    "data": {
                        "session_id": "=session",
                        "move": "=2+2",
                        "player_move": "+SUM(A1:A2)",
                        "fen": "-bad-fen",
                        "actor": "player",
                        "system_status": "@running",
                    },
                },
            )

            source_wb = load_workbook(log_path, data_only=False)
            try:
                pipeline_headers = [cell.value for cell in source_wb["Pipeline_Log"][1]]
                move_col = pipeline_headers.index("move") + 1
                source_col = pipeline_headers.index("source") + 1
                self.assertEqual(source_wb["Pipeline_Log"].cell(row=2, column=move_col).value, "'=2+2")
                self.assertEqual(source_wb["Pipeline_Log"].cell(row=2, column=source_col).value, "'@socket")
            finally:
                source_wb.close()

            exporter.export_session(None, out_path)
            wb = load_workbook(out_path, data_only=False)
            try:
                for ws in wb.worksheets:
                    for row in ws.iter_rows():
                        for cell in row:
                            value = cell.value
                            self.assertNotEqual(cell.data_type, "f", f"{ws.title}!{cell.coordinate} is a formula")
                            if isinstance(value, str):
                                self.assertFalse(
                                    value.lstrip().startswith(("=", "+", "-", "@")),
                                    f"{ws.title}!{cell.coordinate} has unsafe formula-like text: {value!r}",
                                )
                pipeline_headers = [cell.value for cell in wb["Pipeline_Log"][1]]
                move_col = pipeline_headers.index("move") + 1
                source_col = pipeline_headers.index("source") + 1
                self.assertEqual(wb["Pipeline_Log"].cell(row=2, column=move_col).value, "'=2+2")
                self.assertEqual(wb["Pipeline_Log"].cell(row=2, column=source_col).value, "'@socket")
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()

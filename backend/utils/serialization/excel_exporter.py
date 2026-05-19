import json
import os
import queue
import statistics
import threading
import time
import uuid
from collections import Counter
from datetime import datetime
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.utils.logger import logger


FORMULA_PREFIXES = ("=", "+", "-", "@")


def sanitize_excel_cell(value: Any) -> Any:
    """Keep attacker-controlled strings from being interpreted as formulas."""
    if isinstance(value, str) and value.lstrip().startswith(FORMULA_PREFIXES):
        return "'" + value
    return value


class ExcelExporter:
    """
    Research-grade Excel exporter.

    Runtime logging stays append-only in Pipeline_Log. Download/export builds a
    multi-sheet research workbook from that log, so the frontend route remains
    stable while the Excel output becomes useful for experiments and debugging.
    """

    PIPELINE_SHEET = "Pipeline_Log"
    REPORT_SHEETS = [
        "Overview",
        "Game Moves",
        "Vision YOLO",
        "Vision FEN Log",
        "Vision Detections",
        "Vision Mode Comparison",
        "UCCI Trace",
        "Engine AI",
        "Robot Control",
        "System Events",
        "Errors & Warnings",
        "Data Quality",
        "Raw Payload",
    ]

    PIPELINE_BASE_HEADERS = [
        "event_id",
        "session_id",
        "timestamp",
        "event_type",
        "source",
        "trace_id",
        "fen_before",
        "fen_after",
        "ucci_position",
        "move",
        "actor",
        "player_move",
        "ai_move",
        "image_path",
        "detections_count",
        "avg_confidence",
        "min_confidence",
        "yolo_latency_ms",
        "yolo_fps",
        "sahi_enabled",
        "camera_status",
        "engine_score",
        "engine_depth",
        "engine_ms",
        "robot_status",
        "robot_ms",
        "system_status",
        "board_state",
        "raw_payload",
    ]
    PIPELINE_HEADERS = PIPELINE_BASE_HEADERS + ["audit_status", "audit_notes"]

    DATA_QUALITY_HEADERS = [
        "severity",
        "row_number",
        "event_id",
        "field",
        "current_value",
        "suggested_value",
        "reason",
    ]

    EVENT_TYPE_ALIASES = {
        "DIAGNOSTICS.UPDATED": "DIAGNOSTICS_UPDATED",
        "ENGINE.INFO_UPDATED": "ENGINE_INFO_UPDATED",
        "VISION.FRAME_PROCESSED": "VISION_FRAME_PROCESSED",
        "ROBOT.STATUS_UPDATED": "ROBOT_STATUS_UPDATED",
        "GAME.STATE_APPLIED": "GAME_STATE_APPLIED",
    }

    LEGACY_HEADERS = [
        "Timestamp",
        "Session_ID",
        "Event_Type",
        "Source",
        "FEN",
        "Move",
        "Actor",
        "Player_Move",
        "AI_Move",
        "Image",
        "YOLO_Latency_ms",
        "YOLO_FPS",
        "YOLO_Efficiency",
        "YOLO_Confidence",
        "Recognition_Result",
        "Vision_Mode",
        "Vision_Status",
        "Engine_ms",
        "Robot_ms",
        "Engine_Score",
        "Engine_Depth",
        "Robot_Status",
        "System_Status",
        "Raw_Payload",
    ]

    DARK_FILL = PatternFill("solid", fgColor="1F2937")
    RED_FILL = PatternFill("solid", fgColor="FECACA")
    ORANGE_FILL = PatternFill("solid", fgColor="FED7AA")
    GREEN_FILL = PatternFill("solid", fgColor="DCFCE7")
    BLUE_FILL = PatternFill("solid", fgColor="DBEAFE")
    SKY_FILL = PatternFill("solid", fgColor="E0F2FE")
    SUBTLE_FILL = PatternFill("solid", fgColor="F8FAFC")
    SECTION_FILL = PatternFill("solid", fgColor="E5E7EB")
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    THIN_BORDER = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )
    TAB_COLORS = {
        "Overview": "2563EB",
        "Pipeline_Log": "1F2937",
        "Game Moves": "16A34A",
        "Vision YOLO": "0284C7",
        "Vision FEN Log": "0EA5E9",
        "Vision Detections": "38BDF8",
        "Vision Mode Comparison": "7C3AED",
        "UCCI Trace": "6366F1",
        "Engine AI": "F59E0B",
        "Robot Control": "DC2626",
        "System Events": "64748B",
        "Errors & Warnings": "B91C1C",
        "Data Quality": "9333EA",
        "Raw Payload": "475569",
    }

    def __init__(self, filename: str = "chess_robot_experiment.xlsx", subscribe: bool = True):
        self.filename = filename
        self._lock = threading.RLock()
        self._queue = queue.Queue(maxsize=1000)
        self._queue_worker = None
        self._event_counter = 0
        self._subscribed = False
        self._init_workbook()

        if subscribe:
            self.start_subscription()

    def _start_queue_worker(self) -> None:
        if self._queue_worker and self._queue_worker.is_alive():
            return

        def worker():
            while True:
                game_state, event = self._queue.get()
                try:
                    self.log_event(game_state, event)
                finally:
                    self._queue.task_done()

        self._queue_worker = threading.Thread(
            target=worker,
            name="ExcelExporterWriter",
            daemon=True,
        )
        self._queue_worker.start()

    def start_subscription(self) -> None:
        if self._subscribed:
            return
        self._start_queue_worker()
        from backend.events.bus.event_bus import bus

        bus.subscribe_all(self._on_event, key="excel_exporter", replace=True)
        self._subscribed = True

    def _save_and_close(self, wb: Workbook, filename: str) -> None:
        try:
            wb.save(filename)
        finally:
            try:
                wb.close()
            except Exception:
                logger.debug("[ExcelExporter] workbook close failed: %s", filename, exc_info=True)

    def _append_safe(self, ws, values: Iterable[Any]) -> None:
        ws.append([sanitize_excel_cell(value) for value in values])

    def _init_workbook(self) -> None:
        directory = os.path.dirname(os.path.abspath(self.filename))
        if directory:
            os.makedirs(directory, exist_ok=True)

        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = self.PIPELINE_SHEET
            self._append_safe(ws, self.PIPELINE_HEADERS)
            self._save_and_close(wb, self.filename)
            return

        try:
            wb = load_workbook(self.filename)
        except Exception as exc:
            backup = f"{self.filename}.corrupt-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            try:
                os.replace(self.filename, backup)
                logger.warning("[ExcelExporter] corrupt workbook moved to %s: %s", backup, exc)
            except Exception:
                logger.warning("[ExcelExporter] corrupt workbook rebuild in place: %s", exc)
            wb = Workbook()
            ws = wb.active
            ws.title = self.PIPELINE_SHEET
            self._append_safe(ws, self.PIPELINE_HEADERS)
            self._save_and_close(wb, self.filename)
            return

        ws = wb[self.PIPELINE_SHEET] if self.PIPELINE_SHEET in wb.sheetnames else None
        if ws is None:
            ws = wb.create_sheet(self.PIPELINE_SHEET, 0)
            self._append_safe(ws, self.PIPELINE_HEADERS)
            self._save_and_close(wb, self.filename)
            return

        headers = self._sheet_headers(ws)
        if headers == self.PIPELINE_HEADERS:
            wb.close()
            return

        if headers == self.PIPELINE_BASE_HEADERS:
            for header in self.PIPELINE_HEADERS[len(headers):]:
                ws.cell(row=1, column=ws.max_column + 1, value=sanitize_excel_cell(header))
            self._save_and_close(wb, self.filename)
            return

        if headers == self.LEGACY_HEADERS:
            archive_title = self._unique_sheet_title(wb, "Pipeline_Legacy")
            legacy = wb.copy_worksheet(ws)
            legacy.title = archive_title

        ws.delete_rows(1, ws.max_row)
        self._append_safe(ws, self.PIPELINE_HEADERS)
        self._save_and_close(wb, self.filename)

    def _on_event(self, event: Any) -> None:
        from backend.state.store.state_store import state_store

        payload = self._payload(event)
        event_record = {
            "type": self._event_type(event),
            "source": self._source(event),
            "event_id": self._attr_or_dict(event, "event_id", ""),
            "trace_id": self._attr_or_dict(event, "trace_id", payload.get("trace_id", "")),
            "timestamp": self._attr_or_dict(event, "timestamp", payload.get("timestamp", "")),
            "data": payload,
        }
        snapshot = self._state_snapshot(state_store)
        try:
            self._queue.put_nowait((snapshot, event_record))
        except queue.Full:
            logger.warning("[ExcelExporter] event log queue full; dropping %s", event_record.get("type"))

    def _event_type(self, event: Any) -> str:
        value = self._attr_or_dict(event, "event_type", None)
        if value is None and isinstance(event, dict):
            value = event.get("type", "UNKNOWN")
        return value.value if hasattr(value, "value") else str(value or "UNKNOWN")

    def _source(self, event: Any) -> str:
        return str(self._attr_or_dict(event, "source", ""))

    def _payload(self, event: Any) -> Dict[str, Any]:
        if isinstance(event, dict):
            payload = event.get("payload") or event.get("data") or {}
        else:
            payload = getattr(event, "payload", {}) or {}
        return payload if isinstance(payload, dict) else {"value": payload}

    def _attr_or_dict(self, obj: Any, key: str, default: Any = None) -> Any:
        if isinstance(obj, dict):
            return obj.get(key, default)
        return getattr(obj, key, default)

    def _state_snapshot(self, game_state: Any) -> Dict[str, Any]:
        if isinstance(game_state, dict):
            return game_state
        if hasattr(game_state, "to_dict"):
            try:
                snapshot = game_state.to_dict()
                if isinstance(snapshot, dict):
                    return snapshot
            except Exception:
                logger.debug("[ExcelExporter] state_store.to_dict failed", exc_info=True)
        return {}

    def _safe_json(self, value: Any, max_chars: int = 32000) -> str:
        try:
            text = json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            text = str(value)
        if len(text) > max_chars:
            return text[: max_chars - 16] + "...<truncated>"
        return text

    def _load_json(self, value: Any) -> Any:
        if isinstance(value, (dict, list)):
            return value
        if not isinstance(value, str) or not value.strip():
            return {}
        try:
            return json.loads(value)
        except Exception:
            return {}

    def _first_value(self, source: Dict[str, Any], *keys: str, default: Any = "") -> Any:
        if not isinstance(source, dict):
            return default
        for key in keys:
            value = source.get(key)
            if value not in (None, ""):
                return value
        return default

    def _explicit_value(self, source: Dict[str, Any], *keys: str) -> tuple[Any, bool]:
        if not isinstance(source, dict):
            return "", False
        for key in keys:
            if key in source:
                value = source.get(key)
                return "" if value is None else value, True
        return "", False

    def _dict_child(self, source: Dict[str, Any], key: str) -> Dict[str, Any]:
        value = source.get(key) if isinstance(source, dict) else {}
        return value if isinstance(value, dict) else {}

    def _first_from_sources(self, sources: Iterable[Dict[str, Any]], *keys: str, default: Any = "") -> Any:
        for source in sources:
            value = self._first_value(source, *keys, default=None)
            if value not in (None, ""):
                return value
        return default

    def _canonical_event_type(self, value: Any) -> str:
        text = str(value or "UNKNOWN").strip() or "UNKNOWN"
        upper = text.upper()
        return self.EVENT_TYPE_ALIASES.get(upper, upper.replace(".", "_"))

    def _is_uuid(self, value: Any) -> bool:
        try:
            uuid.UUID(str(value))
            return True
        except Exception:
            return False

    def _is_valid_fen(self, value: Any) -> bool:
        if not isinstance(value, str) or not value.strip():
            return False
        try:
            from backend.state.store.validators.fen_validator import FENValidator

            return FENValidator.validate(value)
        except Exception:
            ranks = value.split()[0].split("/")
            return len(ranks) == 10 and all(rank for rank in ranks)

    def _raw_fen_after(self, data: Dict[str, Any]) -> Any:
        value, found = self._explicit_value(data, "fen_after", "after_fen", "fen")
        if found:
            return value
        for child_key in ("game", "board", "vision"):
            value, found = self._explicit_value(self._dict_child(data, child_key), "fen_after", "after_fen", "fen")
            if found:
                return value
        return ""

    def _raw_ucci_position(self, data: Dict[str, Any]) -> Any:
        value, found = self._explicit_value(data, "ucci_position")
        if found:
            return value
        for child_key in ("vision", "engine"):
            value, found = self._explicit_value(self._dict_child(data, child_key), "ucci_position")
            if found:
                return value
        return ""

    def _to_float(self, value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        try:
            return float(str(value).replace("%", "").strip())
        except (TypeError, ValueError):
            return None

    def _timestamp_text(self, value: Any = None) -> str:
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
        if isinstance(value, (int, float)):
            try:
                return datetime.fromtimestamp(float(value)).strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            except Exception:
                logger.debug("[ExcelExporter] invalid numeric timestamp: %r", value, exc_info=True)
        if value:
            return str(value)
        return datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]

    def _next_event_id(self) -> str:
        return str(uuid.uuid4())

    def _extract_detections(self, *sources: Any) -> List[Any]:
        for source in sources:
            if not isinstance(source, dict):
                continue
            detections = self._first_value(source, "detections", "recognition_result", default=None)
            if isinstance(detections, str):
                detections = self._load_json(detections)
            if isinstance(detections, list):
                return detections
        return []

    def _confidence_values(self, detections: Iterable[Any]) -> List[float]:
        values = []
        for det in detections:
            if not isinstance(det, dict):
                continue
            value = self._to_float(
                self._first_value(det, "confidence", "conf", "score", "probability", default=None)
            )
            if value is not None:
                values.append(value)
        return values

    def _event_record(self, game_state: Any, event: Dict[str, Any]) -> Dict[str, Any]:
        snapshot = self._state_snapshot(game_state)
        game = snapshot.get("game", {}) if isinstance(snapshot.get("game"), dict) else {}

        data = event.get("data", {}) if isinstance(event.get("data"), dict) else {}
        event_type_raw = str(event.get("type", "UNKNOWN"))
        event_type = self._canonical_event_type(event_type_raw)
        upper_type = event_type.upper()
        vision_payload = self._dict_child(data, "vision")
        engine_payload = self._dict_child(data, "engine")
        robot_payload = self._dict_child(data, "robot")
        system_payload = self._dict_child(data, "system")
        game_payload = self._dict_child(data, "game")
        data_sources = [data]

        vision_context = "VISION" in upper_type or bool(vision_payload) or any(
            key in data for key in ("detections", "detections_count", "latency_ms", "yolo_latency_ms", "camera_status", "vision_status")
        )
        engine_context = "ENGINE" in upper_type or any(
            key in data for key in ("score", "engine_score", "depth", "engine_depth", "best_move", "bestmove", "ai_move")
        )
        robot_context = "ROBOT" in upper_type or any(
            key in data for key in ("robot_status", "robot_ms")
        )

        detections = self._extract_detections(data, vision_payload)
        confidences = self._confidence_values(detections)
        avg_confidence = self._first_from_sources([data, vision_payload], "avg_confidence", "confidence", default="")
        if avg_confidence == "" and confidences:
            avg_confidence = round(sum(confidences) / len(confidences), 4)
        min_confidence = self._first_from_sources([data, vision_payload], "min_confidence", default="")
        if min_confidence == "" and confidences:
            min_confidence = round(min(confidences), 4)

        fen_after = self._raw_fen_after(data)
        fen_before = self._first_value(data, "fen_before", "before_fen", default="")
        ucci_position = self._raw_ucci_position(data)

        move = self._first_from_sources([data, engine_payload], "move", "uci", "best_move", "bestmove", "ai_move", "player_move", default="")
        actor = self._first_value(data, "actor", "side", "turn", default="")
        if not actor:
            actor = self._infer_actor(event_type)

        player_move = self._first_value(data, "player_move", default="")
        ai_move = self._first_from_sources([data, engine_payload], "ai_move", "best_move", "bestmove", default="")
        if not player_move and actor in ("player", "red", "black"):
            player_move = move
        if not ai_move and (actor in ("ai", "engine") or engine_context):
            ai_move = move

        board_state = self._first_from_sources([data, vision_payload], "board_state", "stable_state", default={})
        detections_count = self._first_from_sources([data, vision_payload], "detections_count", default="")
        if detections_count == "" and (detections or vision_context):
            detections_count = len(detections)

        camera_status = self._first_from_sources([data, vision_payload], "camera_status", "vision_status", default="")
        if not camera_status and vision_payload:
            camera_status = self._first_value(vision_payload, "status", default="")
        if not camera_status and vision_context:
            camera_status = self._first_value(data, "status", default="")

        engine_score = ""
        engine_depth = ""
        engine_ms = ""
        if engine_context:
            engine_score = self._first_from_sources([data, engine_payload], "score", "engine_score", "eval", default="")
            engine_depth = self._first_from_sources([data, engine_payload], "depth", "engine_depth", default="")
            engine_ms = self._first_from_sources([data, engine_payload], "engine_ms", "decision_time_ms", "time", default="")

        robot_status = self._first_from_sources([data, robot_payload], "robot_status", default="")
        if not robot_status and robot_payload:
            robot_status = self._first_value(robot_payload, "status", default="")
        if not robot_status and robot_context:
            robot_status = self._first_value(data, "status", default="")
        robot_ms = self._first_from_sources([data, robot_payload], "robot_ms", "duration_ms", default="") if robot_context else ""

        event_id_raw = event.get("event_id") or data.get("event_id")
        session_id = data.get("session_id") or event.get("session_id") or snapshot.get("session_id") or game.get("session_id") or ""
        trace_id = event.get("trace_id") or data.get("trace_id") or ""

        record = {
            "event_id": event_id_raw or self._next_event_id(),
            "session_id": session_id,
            "timestamp": self._timestamp_text(event.get("timestamp") or data.get("timestamp")),
            "event_type": event_type,
            "source": event.get("source", ""),
            "trace_id": trace_id,
            "fen_before": fen_before,
            "fen_after": fen_after,
            "ucci_position": ucci_position,
            "move": move,
            "actor": actor,
            "player_move": player_move,
            "ai_move": ai_move,
            "image_path": self._first_from_sources([data, vision_payload], "image_path", "frame_path", "image", default=""),
            "detections_count": detections_count,
            "avg_confidence": avg_confidence if vision_context else "",
            "min_confidence": min_confidence if vision_context else "",
            "yolo_latency_ms": self._first_from_sources([data, vision_payload], "yolo_latency_ms", "latency_ms", default="") if vision_context else "",
            "yolo_fps": self._first_from_sources([data, vision_payload], "fps", "yolo_fps", default="") if vision_context else "",
            "sahi_enabled": self._first_from_sources([data, vision_payload], "sahi_enabled", default="") if vision_context else "",
            "camera_status": camera_status,
            "engine_score": engine_score,
            "engine_depth": engine_depth,
            "engine_ms": engine_ms,
            "robot_status": robot_status,
            "robot_ms": robot_ms,
            "system_status": self._first_from_sources([data, system_payload, game_payload], "system_status", "status", "phase", default=""),
            "board_state": self._safe_json(board_state),
            "raw_payload": self._safe_json(data),
            "_event_type_raw": event_type_raw,
            "_event_id_generated": not bool(event_id_raw),
        }
        self._apply_audit(record)
        return record

    def _values_equal(self, left: Any, right: Any) -> bool:
        if left in (None, "") and right in (None, ""):
            return True
        left_num = self._to_float(left)
        right_num = self._to_float(right)
        if left_num is not None and right_num is not None:
            return abs(left_num - right_num) < 1e-9
        return str(left) == str(right)

    def _quality_row(
        self,
        record: Dict[str, Any],
        severity: str,
        field: str,
        current_value: Any,
        suggested_value: Any,
        reason: str,
    ) -> Dict[str, Any]:
        return {
            "severity": severity,
            "row_number": record.get("_row_number", ""),
            "event_id": record.get("event_id", ""),
            "field": field,
            "current_value": "" if current_value is None else current_value,
            "suggested_value": "" if suggested_value is None else suggested_value,
            "reason": reason,
        }

    def _raw_field_suggestion(self, record: Dict[str, Any], raw: Dict[str, Any], field: str) -> tuple[Any, bool]:
        event_type = str(record.get("event_type", "")).upper()
        vision_payload = self._dict_child(raw, "vision")
        engine_payload = self._dict_child(raw, "engine")
        robot_payload = self._dict_child(raw, "robot")

        if field == "camera_status":
            value = self._first_from_sources([raw, vision_payload], "camera_status", "vision_status", default="")
            if value in (None, ""):
                value = self._first_value(vision_payload, "status", default="")
            if value in (None, "") and "VISION" in event_type:
                value = self._first_value(raw, "status", default="")
            return value, value not in (None, "")

        if field == "engine_score":
            if "ENGINE" not in event_type:
                return "", False
            value = self._first_from_sources([raw, engine_payload], "score", "engine_score", "eval", default="")
            return value, value not in (None, "")

        if field == "engine_depth":
            if "ENGINE" not in event_type:
                return "", False
            value = self._first_from_sources([raw, engine_payload], "depth", "engine_depth", default="")
            return value, value not in (None, "")

        if field == "ai_move":
            if "ENGINE" not in event_type:
                return "", False
            value = self._first_from_sources([raw, engine_payload], "ai_move", "best_move", "bestmove", "move", default="")
            return value, value not in (None, "")

        if field == "fen_after":
            value, found = self._explicit_value(raw, "fen_after", "after_fen", "fen")
            if found:
                return value, True
            value = self._raw_fen_after(raw)
            return value, value not in (None, "")

        if field == "ucci_position":
            value = self._raw_ucci_position(raw)
            return value, value not in (None, "")

        return "", False

    def _quality_issues(self, record: Dict[str, Any]) -> List[Dict[str, Any]]:
        issues: List[Dict[str, Any]] = []
        raw = self._load_json(record.get("raw_payload"))
        raw = raw if isinstance(raw, dict) else {}
        event_type = str(record.get("event_type", "") or "")
        raw_event_type = str(record.get("_event_type_raw") or event_type)
        canonical = self._canonical_event_type(raw_event_type)

        generated_note = "missing event_id; generated UUID" in str(record.get("audit_notes", ""))
        if record.get("_event_id_generated") or generated_note:
            issues.append(
                self._quality_row(
                    record,
                    "Info",
                    "event_id",
                    "",
                    record.get("event_id", ""),
                    "missing event_id; generated UUID",
                )
            )
        elif record.get("event_id") and not self._is_uuid(record.get("event_id")):
            issues.append(
                self._quality_row(
                    record,
                    "Warning",
                    "event_id",
                    record.get("event_id", ""),
                    "",
                    "event_id is not a UUID",
                )
            )

        if record.get("session_id") in (None, "", "N/A"):
            issues.append(self._quality_row(record, "Warning", "session_id", record.get("session_id", ""), "", "missing session_id"))

        if record.get("trace_id") in (None, ""):
            issues.append(self._quality_row(record, "Warning", "trace_id", record.get("trace_id", ""), "", "missing trace_id"))

        canonicalized_note = "event_type: event type canonicalized" in str(record.get("audit_notes", ""))
        if raw_event_type != canonical or "." in raw_event_type or canonicalized_note:
            issues.append(
                self._quality_row(
                    record,
                    "Info",
                    "event_type",
                    raw_event_type if not canonicalized_note else record.get("event_type", ""),
                    canonical,
                    "event type canonicalized",
                )
            )

        fen_after = record.get("fen_after")
        if fen_after not in (None, "") and not self._is_valid_fen(fen_after):
            issues.append(self._quality_row(record, "Warning", "fen_after", fen_after, "", "invalid or non-standard Xiangqi FEN"))

        for field in ("camera_status", "engine_score", "engine_depth", "ai_move", "fen_after", "ucci_position"):
            suggested, has_suggestion = self._raw_field_suggestion(record, raw, field)
            if has_suggestion and not self._values_equal(record.get(field, ""), suggested):
                issues.append(
                    self._quality_row(
                        record,
                        "Warning",
                        field,
                        record.get(field, ""),
                        suggested,
                        "value differs from raw_payload",
                    )
                )

        upper_type = event_type.upper()
        non_applicable_zero_fields = []
        if "ENGINE" not in upper_type:
            non_applicable_zero_fields.extend(["engine_score", "engine_depth", "engine_ms"])
        if "VISION" not in upper_type:
            non_applicable_zero_fields.extend(["detections_count", "avg_confidence", "min_confidence", "yolo_latency_ms", "yolo_fps"])
        if "ROBOT" not in upper_type:
            non_applicable_zero_fields.extend(["robot_ms"])

        for field in non_applicable_zero_fields:
            if record.get(field) in (0, 0.0, "0", "0.0"):
                issues.append(
                    self._quality_row(
                        record,
                        "Warning",
                        field,
                        record.get(field),
                        "",
                        "zero appears to represent a non-applicable metric",
                    )
                )

        return issues

    def _apply_audit(self, record: Dict[str, Any]) -> None:
        issues = self._quality_issues(record)
        record["audit_status"] = "ok" if not issues else "review"
        record["audit_notes"] = "; ".join(
            f"{issue['field']}: {issue['reason']}" for issue in issues
        )[:32000]

    def _data_quality_rows(self, records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        for record in records:
            rows.extend(self._quality_issues(record))
        return rows

    def _infer_actor(self, event_type: str) -> str:
        upper = event_type.upper()
        if "ENGINE" in upper:
            return "ai"
        if "ROBOT" in upper:
            return "robot"
        if "VISION" in upper:
            return "vision"
        if "PLAYER" in upper:
            return "player"
        return ""

    def log_event(self, game_state: Any, event: Dict[str, Any]) -> None:
        """Append one normalized runtime event into Pipeline_Log."""
        try:
            with self._lock:
                wb = load_workbook(self.filename)
                try:
                    ws = wb[self.PIPELINE_SHEET]
                    record = self._event_record(game_state, event)
                    self._append_safe(ws, [record.get(header, "") for header in self.PIPELINE_HEADERS])
                    wb.save(self.filename)
                finally:
                    wb.close()
        except Exception as exc:
            logger.error("[ExcelExporter] log_event failed: %s", exc, exc_info=True)

    def export_session(self, session_id: Optional[str], target_filename: str) -> None:
        """Build a multi-sheet research workbook for one session or all sessions."""
        try:
            self._flush_queue(timeout=2.0)
            with self._lock:
                records = self._load_records(session_id)
                wb = Workbook()
                default = wb.active
                wb.remove(default)

                self._write_overview(wb, records, session_id)
                self._write_pipeline(wb, records)
                self._write_game_moves(wb, records)
                self._write_vision_yolo(wb, records)
                self._write_vision_fen(wb, records)
                self._write_vision_detections(wb, records)
                self._write_vision_mode_comparison(wb, records)
                self._write_ucci_trace(wb, records)
                self._write_engine_ai(wb, records)
                self._write_robot_control(wb, records)
                self._write_system_events(wb, records)
                self._write_errors(wb, records)
                self._write_data_quality(wb, records)
                self._write_raw_payload(wb, records)

                self._style_workbook(wb)
                self._save_and_close(wb, target_filename)
        except Exception as exc:
            logger.error("[ExcelExporter] export_session failed: %s", exc, exc_info=True)
            raise

    def _flush_queue(self, timeout: float = 2.0) -> None:
        deadline = time.time() + timeout
        while getattr(self._queue, "unfinished_tasks", 0) > 0 and time.time() < deadline:
            time.sleep(0.02)

    def _load_records(self, session_id: Optional[str]) -> List[Dict[str, Any]]:
        if not os.path.exists(self.filename):
            self._init_workbook()

        records = []
        wb = load_workbook(self.filename, data_only=True)
        try:
            if self.PIPELINE_SHEET in wb.sheetnames:
                records.extend(self._records_from_sheet(wb[self.PIPELINE_SHEET], self.PIPELINE_HEADERS))
            for sheet_name in wb.sheetnames:
                if sheet_name.startswith("Pipeline_Legacy"):
                    records.extend(self._legacy_records(wb[sheet_name]))
        finally:
            wb.close()

        if session_id:
            return [row for row in records if str(row.get("session_id", "")) == str(session_id)]
        return records

    def _records_from_sheet(self, ws: Any, headers: List[str]) -> List[Dict[str, Any]]:
        actual = self._sheet_headers(ws)
        if actual not in (headers, self.PIPELINE_BASE_HEADERS):
            return []
        rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            record = dict(zip(actual, row))
            if any(value not in (None, "") for value in record.values()):
                normalized = {key: "" if value is None else value for key, value in record.items()}
                for header in self.PIPELINE_HEADERS:
                    normalized.setdefault(header, "")
                normalized["_row_number"] = row_idx
                normalized["_event_type_raw"] = normalized.get("event_type", "")
                self._apply_audit(normalized)
                rows.append(normalized)
        return rows

    def _legacy_records(self, ws: Any) -> List[Dict[str, Any]]:
        if self._sheet_headers(ws) != self.LEGACY_HEADERS:
            return []
        output = []
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            legacy = dict(zip(self.LEGACY_HEADERS, row))
            raw = legacy.get("Raw_Payload") or "{}"
            recognition = self._load_json(legacy.get("Recognition_Result"))
            detections_count = len(recognition) if isinstance(recognition, list) else ""
            fen = legacy.get("FEN") or ""
            record = {
                    "event_id": str(uuid.uuid4()),
                    "session_id": legacy.get("Session_ID") or "",
                    "timestamp": legacy.get("Timestamp") or "",
                    "event_type": self._canonical_event_type(legacy.get("Event_Type") or "UNKNOWN"),
                    "source": legacy.get("Source") or "",
                    "trace_id": "",
                    "fen_before": "",
                    "fen_after": fen,
                    "ucci_position": "",
                    "move": legacy.get("Move") or "",
                    "actor": legacy.get("Actor") or "",
                    "player_move": legacy.get("Player_Move") or "",
                    "ai_move": legacy.get("AI_Move") or "",
                    "image_path": legacy.get("Image") or "",
                    "detections_count": detections_count,
                    "avg_confidence": legacy.get("YOLO_Confidence") or "",
                    "min_confidence": legacy.get("YOLO_Confidence") or "",
                    "yolo_latency_ms": legacy.get("YOLO_Latency_ms") or "",
                    "yolo_fps": legacy.get("YOLO_FPS") or "",
                    "sahi_enabled": "",
                    "camera_status": legacy.get("Vision_Status") or legacy.get("Vision_Mode") or "",
                    "engine_score": legacy.get("Engine_Score") or "",
                    "engine_depth": legacy.get("Engine_Depth") or "",
                    "engine_ms": legacy.get("Engine_ms") or "",
                    "robot_status": legacy.get("Robot_Status") or "",
                    "robot_ms": legacy.get("Robot_ms") or "",
                    "system_status": legacy.get("System_Status") or "",
                    "board_state": "",
                    "raw_payload": raw,
                    "_event_type_raw": legacy.get("Event_Type") or "UNKNOWN",
                    "_event_id_generated": True,
                }
            self._apply_audit(record)
            output.append(record)
        return output

    def _write_overview(self, wb: Workbook, records: List[Dict[str, Any]], session_id: Optional[str]) -> None:
        ws = wb.create_sheet("Overview")
        ws["A1"] = "S.M.A.R.T Chess Robot Research Report"
        ws["A1"].font = Font(bold=True, size=14, color="111827")
        self._append_safe(ws, [])

        vision_records = [row for row in records if self._is_vision(row)]
        engine_records = [row for row in records if self._is_engine(row)]
        robot_records = [row for row in records if self._is_robot(row)]
        move_records = [row for row in records if self._is_move(row)]
        warnings = self._warning_rows(records)

        self._append_safe(ws, ["Metric", "Value"])
        metrics = [
            ("Generated At", self._timestamp_text()),
            ("Session Filter", session_id or "all"),
            ("Total Events", len(records)),
            ("Game Moves", len(move_records)),
            ("Vision Events", len(vision_records)),
            ("YOLO Avg Confidence", self._average(records, "avg_confidence")),
            ("YOLO Min Confidence", self._minimum(records, "min_confidence")),
            ("YOLO Avg Latency ms", self._average(records, "yolo_latency_ms")),
            ("AI Avg Decision ms", self._average(engine_records, "engine_ms")),
            ("Robot Avg Execution ms", self._average(robot_records, "robot_ms")),
            ("Errors / Warnings", len(warnings)),
        ]
        for label, value in metrics:
            self._append_safe(ws, [label, value])

        self._append_safe(ws, [])
        self._append_safe(ws, ["Event Type", "Count"])
        for event_type, count in Counter(row.get("event_type", "") for row in records).most_common():
            self._append_safe(ws, [event_type, count])

    def _write_pipeline(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        self._write_rows(wb, self.PIPELINE_SHEET, self.PIPELINE_HEADERS, records)

    def _write_game_moves(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["event_id", "timestamp", "actor", "move", "player_move", "ai_move", "fen_before", "fen_after", "system_status", "trace_id"]
        self._write_rows(wb, "Game Moves", headers, [row for row in records if self._is_move(row)])

    def _write_vision_yolo(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = [
            "event_id",
            "timestamp",
            "detections_count",
            "avg_confidence",
            "min_confidence",
            "yolo_latency_ms",
            "yolo_fps",
            "sahi_enabled",
            "camera_status",
            "fen_after",
            "image_path",
            "trace_id",
        ]
        self._write_rows(wb, "Vision YOLO", headers, [row for row in records if self._is_vision(row)])

    def _write_vision_fen(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = [
            "event_id",
            "timestamp",
            "fen_after",
            "ucci_position",
            "detections_count",
            "avg_confidence",
            "min_confidence",
            "yolo_latency_ms",
            "board_state",
            "trace_id",
        ]
        rows = [row for row in records if self._is_vision_fen(row)]
        self._write_rows(wb, "Vision FEN Log", headers, rows)

    def _write_vision_detections(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = [
            "event_id",
            "timestamp",
            "class_name",
            "confidence",
            "bbox",
            "cell",
            "fen_after",
            "yolo_latency_ms",
            "image_path",
            "trace_id",
        ]
        rows = []
        for record in records:
            if not self._is_vision(record):
                continue
            payload = self._load_json(record.get("raw_payload"))
            detections = self._extract_detections(payload)
            if not detections:
                rows.append(
                    {
                        "event_id": record.get("event_id", ""),
                        "timestamp": record.get("timestamp", ""),
                        "class_name": "",
                        "confidence": "",
                        "bbox": "",
                        "cell": "",
                        "fen_after": record.get("fen_after", ""),
                        "yolo_latency_ms": record.get("yolo_latency_ms", ""),
                        "image_path": record.get("image_path", ""),
                        "trace_id": record.get("trace_id", ""),
                    }
                )
                continue
            for det in detections:
                rows.append(self._detection_row(record, det))
        self._write_rows(wb, "Vision Detections", headers, rows)

    def _write_vision_mode_comparison(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = [
            "row_type",
            "mode",
            "frame_id",
            "frames",
            "ok_frames",
            "status",
            "skip_reason",
            "avg_fps",
            "fps",
            "avg_inference_latency_ms",
            "inference_latency_ms",
            "avg_end_to_end_latency_ms",
            "end_to_end_latency_ms",
            "p95_end_to_end_latency_ms",
            "fen",
            "avg_detections",
            "detections_count",
            "avg_small_object_rate",
            "small_object_rate",
            "small_object_count",
            "fen_valid_rate",
            "fen_valid",
            "stable_update_rate",
            "stable_update",
            "roi_applied",
            "roi",
            "map_50",
            "recall",
            "metric_note",
            "requires_annotations",
            "detections_json",
            "board_state_json",
            "trace_id",
        ]
        benchmark_rows = self._benchmark_payloads(records)
        output = self._benchmark_summary_rows(benchmark_rows)
        for row in benchmark_rows:
            detail = {"row_type": "detail"}
            detail.update(row)
            output.append(detail)
        self._write_rows(wb, "Vision Mode Comparison", headers, output)

    def _benchmark_payloads(self, records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows = []
        for record in records:
            if str(record.get("event_type", "")).upper() != "VISION_BENCHMARK_RESULT":
                continue
            raw = self._load_json(record.get("raw_payload"))
            if not isinstance(raw, dict):
                continue
            row = dict(raw)
            row.setdefault("trace_id", record.get("trace_id", ""))
            row.setdefault("mode", "")
            row.setdefault("map_50", "N/A")
            row.setdefault("recall", "N/A")
            row.setdefault("metric_note", "requires_annotations")
            row.setdefault("requires_annotations", True)
            rows.append(row)
        return rows

    def _benchmark_summary_rows(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        grouped: Dict[str, List[Dict[str, Any]]] = {}
        for row in rows:
            grouped.setdefault(str(row.get("mode", "")), []).append(row)

        output = []
        for mode in sorted(grouped):
            items = grouped[mode]
            ok_items = [item for item in items if str(item.get("status", "")).lower() == "ok"]
            source = ok_items or items
            output.append(
                {
                    "row_type": "summary",
                    "mode": mode,
                    "frames": len(items),
                    "ok_frames": len(ok_items),
                    "avg_fps": self._average_dicts(source, "fps"),
                    "avg_inference_latency_ms": self._average_dicts(source, "inference_latency_ms"),
                    "avg_end_to_end_latency_ms": self._average_dicts(source, "end_to_end_latency_ms"),
                    "p95_end_to_end_latency_ms": self._p95_dicts(source, "end_to_end_latency_ms"),
                    "avg_detections": self._average_dicts(source, "detections_count"),
                    "avg_small_object_rate": self._average_dicts(source, "small_object_rate"),
                    "fen_valid_rate": self._bool_rate(source, "fen_valid"),
                    "stable_update_rate": self._bool_rate(source, "stable_update"),
                    "map_50": "N/A",
                    "recall": "N/A",
                    "metric_note": "requires_annotations",
                    "requires_annotations": True,
                }
            )
        return output

    def _write_ucci_trace(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["event_id", "timestamp", "ucci_position", "ai_move", "engine_score", "engine_depth", "engine_ms", "fen_after", "trace_id"]
        rows = []
        for row in records:
            if row.get("ucci_position") or self._is_engine(row) or row.get("ai_move"):
                rows.append(row)
        self._write_rows(wb, "UCCI Trace", headers, rows)

    def _write_engine_ai(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["event_id", "timestamp", "ai_move", "move", "engine_score", "engine_depth", "engine_ms", "fen_after", "trace_id", "raw_payload"]
        self._write_rows(wb, "Engine AI", headers, [row for row in records if self._is_engine(row)])

    def _write_robot_control(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["event_id", "timestamp", "robot_status", "move", "robot_ms", "system_status", "trace_id", "raw_payload"]
        self._write_rows(wb, "Robot Control", headers, [row for row in records if self._is_robot(row)])

    def _write_system_events(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["event_id", "timestamp", "event_type", "source", "session_id", "trace_id", "system_status", "move", "fen_after"]
        self._write_rows(wb, "System Events", headers, records)

    def _write_errors(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["severity", "reason", "event_id", "timestamp", "event_type", "source", "metric", "value", "raw_payload"]
        self._write_rows(wb, "Errors & Warnings", headers, self._warning_rows(records))

    def _write_data_quality(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        self._write_rows(wb, "Data Quality", self.DATA_QUALITY_HEADERS, self._data_quality_rows(records))

    def _write_raw_payload(self, wb: Workbook, records: List[Dict[str, Any]]) -> None:
        headers = ["event_id", "timestamp", "event_type", "source", "trace_id", "session_id", "raw_payload"]
        self._write_rows(wb, "Raw Payload", headers, records)

    def _write_rows(self, wb: Workbook, title: str, headers: List[str], rows: List[Dict[str, Any]]) -> None:
        ws = wb.create_sheet(title)
        self._append_safe(ws, headers)
        for row in rows:
            self._append_safe(ws, [row.get(header, "") for header in headers])

    def _detection_row(self, record: Dict[str, Any], det: Any) -> Dict[str, Any]:
        det = det if isinstance(det, dict) else {}
        class_name = self._first_value(det, "class_name", "class", "label", "name", default="")
        confidence = self._first_value(det, "confidence", "conf", "score", default="")
        bbox = self._first_value(det, "bbox", "box", "xyxy", default="")
        cell = self._first_value(det, "cell", "board_cell", "grid", "square", default="")
        return {
            "event_id": record.get("event_id", ""),
            "timestamp": record.get("timestamp", ""),
            "class_name": class_name,
            "confidence": confidence,
            "bbox": self._safe_json(bbox),
            "cell": self._safe_json(cell) if isinstance(cell, (dict, list)) else cell,
            "fen_after": record.get("fen_after", ""),
            "yolo_latency_ms": record.get("yolo_latency_ms", ""),
            "image_path": record.get("image_path", ""),
            "trace_id": record.get("trace_id", ""),
        }

    def _warning_rows(self, records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        rows = []
        for record in records:
            upper_type = str(record.get("event_type", "")).upper()
            raw = self._load_json(record.get("raw_payload"))
            error_text = ""
            if isinstance(raw, dict):
                error_text = self._first_value(raw, "error", "exception", "message", default="")

            if "ERROR" in upper_type or "WARNING" in upper_type or error_text:
                rows.append(self._warning_row(record, "Critical" if "ERROR" in upper_type else "Warning", error_text or upper_type, "event", record.get("event_type", "")))

            min_conf = self._to_float(record.get("min_confidence") or record.get("avg_confidence"))
            if min_conf is not None and min_conf < 0.5:
                rows.append(self._warning_row(record, "Warning", "Low YOLO confidence", "confidence", min_conf))

            latency = self._to_float(record.get("yolo_latency_ms"))
            if latency is not None and latency > 1000:
                rows.append(self._warning_row(record, "Warning", "High YOLO latency", "yolo_latency_ms", latency))

            engine_ms = self._to_float(record.get("engine_ms"))
            if engine_ms is not None and engine_ms > 5000:
                rows.append(self._warning_row(record, "Warning", "High engine decision time", "engine_ms", engine_ms))

            robot_status = str(record.get("robot_status", "")).lower()
            if robot_status in ("error", "fault", "offline", "disconnected"):
                rows.append(self._warning_row(record, "Critical", "Robot abnormal status", "robot_status", robot_status))
        return rows

    def _warning_row(self, record: Dict[str, Any], severity: str, reason: str, metric: str, value: Any) -> Dict[str, Any]:
        return {
            "severity": severity,
            "reason": reason,
            "event_id": record.get("event_id", ""),
            "timestamp": record.get("timestamp", ""),
            "event_type": record.get("event_type", ""),
            "source": record.get("source", ""),
            "metric": metric,
            "value": value,
            "raw_payload": record.get("raw_payload", ""),
        }

    def _is_vision(self, row: Dict[str, Any]) -> bool:
        event_type = str(row.get("event_type", "")).upper()
        if event_type == "VISION_BENCHMARK_RESULT":
            return False
        return "VISION" in event_type or row.get("detections_count") not in ("", None) or row.get("yolo_latency_ms") not in ("", None)

    def _is_vision_fen(self, row: Dict[str, Any]) -> bool:
        event_type = str(row.get("event_type", "")).upper()
        if event_type == "VISION_BENCHMARK_RESULT":
            return False
        return "VISION" in event_type and bool(row.get("fen_after"))

    def _is_engine(self, row: Dict[str, Any]) -> bool:
        event_type = str(row.get("event_type", "")).upper()
        return "ENGINE" in event_type or row.get("engine_ms") not in ("", None) or row.get("ai_move") not in ("", None)

    def _is_robot(self, row: Dict[str, Any]) -> bool:
        event_type = str(row.get("event_type", "")).upper()
        return "ROBOT" in event_type or row.get("robot_status") not in ("", None) or row.get("robot_ms") not in ("", None)

    def _is_move(self, row: Dict[str, Any]) -> bool:
        event_type = str(row.get("event_type", "")).upper()
        return "MOVE" in event_type or bool(row.get("move") or row.get("player_move") or row.get("ai_move"))

    def _average(self, records: List[Dict[str, Any]], key: str) -> Any:
        values = [self._to_float(row.get(key)) for row in records]
        values = [value for value in values if value is not None]
        if not values:
            return ""
        return round(sum(values) / len(values), 3)

    def _average_dicts(self, rows: List[Dict[str, Any]], key: str) -> Any:
        values = [self._to_float(row.get(key)) for row in rows]
        values = [value for value in values if value is not None]
        if not values:
            return ""
        return round(sum(values) / len(values), 3)

    def _p95_dicts(self, rows: List[Dict[str, Any]], key: str) -> Any:
        values = sorted(value for value in (self._to_float(row.get(key)) for row in rows) if value is not None)
        if not values:
            return ""
        if len(values) == 1:
            return round(values[0], 3)
        try:
            return round(statistics.quantiles(values, n=20, method="inclusive")[18], 3)
        except Exception:
            index = min(len(values) - 1, int(round(0.95 * (len(values) - 1))))
            return round(values[index], 3)

    def _bool_rate(self, rows: List[Dict[str, Any]], key: str) -> Any:
        if not rows:
            return ""
        hits = 0
        for row in rows:
            value = row.get(key)
            if isinstance(value, str):
                value = value.strip().lower() in ("1", "true", "yes", "ok")
            if bool(value):
                hits += 1
        return round(hits / len(rows), 4)

    def _minimum(self, records: List[Dict[str, Any]], key: str) -> Any:
        values = [self._to_float(row.get(key)) for row in records]
        values = [value for value in values if value is not None]
        if not values:
            return ""
        return round(min(values), 3)

    def _style_workbook(self, wb: Workbook) -> None:
        for ws in wb.worksheets:
            self._style_sheet(ws)
            self._style_sheet_tab(ws)
            if ws.title == "Overview":
                self._style_overview(ws)
            if ws.title == "Vision Mode Comparison":
                self._style_vision_comparison(ws)
        self._apply_report_conditionals(wb)

    def _style_sheet(self, ws: Any) -> None:
        header_row = self._header_row_index(ws)
        if header_row is None:
            return

        ws.sheet_view.showGridLines = False
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        ws.page_margins.left = 0.35
        ws.page_margins.right = 0.35
        ws.page_margins.top = 0.5
        ws.page_margins.bottom = 0.5

        for cell in ws[header_row]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = self.DARK_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = self.THIN_BORDER
        ws.row_dimensions[header_row].height = 24

        ws.freeze_panes = f"A{header_row + 1}"
        if ws.max_row >= header_row:
            ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ws.max_column)}{ws.max_row}"

        for col_idx, column_cells in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in column_cells:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = self.THIN_BORDER
                if cell.row > header_row:
                    cell.fill = self.SUBTLE_FILL if (cell.row - header_row) % 2 == 0 else self.WHITE_FILL
            header = str(ws.cell(row=header_row, column=col_idx).value or "")
            limit = 95 if header in ("raw_payload", "detections_json", "board_state_json") else 60
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), limit)
            self._apply_column_number_format(ws, header, col_idx, header_row)

    def _style_sheet_tab(self, ws: Any) -> None:
        color = self.TAB_COLORS.get(ws.title)
        if color:
            ws.sheet_properties.tabColor = color

    def _style_overview(self, ws: Any) -> None:
        if ws.max_row < 1:
            return
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, ws.max_column))
        title = ws.cell(row=1, column=1)
        title.fill = PatternFill("solid", fgColor="EFF6FF")
        title.font = Font(bold=True, size=16, color="1E3A8A")
        title.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 30

        for row in range(1, ws.max_row + 1):
            label = ws.cell(row=row, column=1).value
            value = ws.cell(row=row, column=2).value
            if label == "Metric" and value == "Value":
                continue
            if label and value not in (None, "") and row > 3:
                ws.cell(row=row, column=1).font = Font(bold=True, color="334155")
                ws.cell(row=row, column=2).font = Font(bold=True, color="0F172A")
            if label == "Event Type":
                for cell in ws[row]:
                    cell.fill = self.SECTION_FILL
                    cell.font = Font(bold=True, color="111827")

    def _style_vision_comparison(self, ws: Any) -> None:
        header_row = self._header_row_index(ws)
        if header_row is None:
            return
        row_type_col = self._column_by_header(ws, "row_type")
        if not row_type_col:
            return
        for row in range(header_row + 1, ws.max_row + 1):
            row_type = str(ws.cell(row=row, column=row_type_col).value or "").lower()
            if row_type == "summary":
                for cell in ws[row]:
                    cell.fill = self.BLUE_FILL
                    cell.font = Font(bold=True, color="1E3A8A")

    def _apply_column_number_format(self, ws: Any, header: str, col_idx: int, header_row: int) -> None:
        if ws.max_row <= header_row:
            return
        lowered = header.lower()
        if lowered in ("fps", "avg_fps", "small_object_rate", "avg_small_object_rate", "fen_valid_rate", "stable_update_rate"):
            fmt = "0.00"
        elif lowered.endswith("_ms") or "latency" in lowered or lowered in ("robot_ms", "engine_ms", "yolo_latency_ms"):
            fmt = "0.0"
        elif lowered in ("detections_count", "small_object_count", "frames", "ok_frames", "frame_id", "engine_depth"):
            fmt = "0"
        elif "confidence" in lowered or lowered == "engine_score":
            fmt = "0.000"
        else:
            return
        for row in range(header_row + 1, ws.max_row + 1):
            ws.cell(row=row, column=col_idx).number_format = fmt

    def _apply_report_conditionals(self, wb: Workbook) -> None:
        for sheet_name in ("Vision YOLO", "Vision FEN Log", "Vision Detections"):
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            self._numeric_rule(ws, "confidence", "lessThan", 0.5, self.RED_FILL)
            self._numeric_rule(ws, "min_confidence", "lessThan", 0.5, self.RED_FILL)
            self._numeric_rule(ws, "avg_confidence", "lessThan", 0.5, self.RED_FILL)
            self._numeric_rule(ws, "yolo_latency_ms", "greaterThan", 1000, self.ORANGE_FILL)

        if "Errors & Warnings" in wb.sheetnames:
            ws = wb["Errors & Warnings"]
            if ws.max_row < 2:
                return
            severity_col = self._column_by_header(ws, "severity")
            if severity_col:
                letter = get_column_letter(severity_col)
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="equal", formula=['"Critical"'], fill=self.RED_FILL),
                )
                ws.conditional_formatting.add(
                    f"{letter}2:{letter}{ws.max_row}",
                    CellIsRule(operator="equal", formula=['"Warning"'], fill=self.ORANGE_FILL),
                )

    def _numeric_rule(self, ws: Any, header: str, operator: str, threshold: float, fill: PatternFill) -> None:
        column = self._column_by_header(ws, header)
        if not column or ws.max_row < 2:
            return
        letter = get_column_letter(column)
        ws.conditional_formatting.add(
            f"{letter}2:{letter}{ws.max_row}",
            CellIsRule(operator=operator, formula=[str(threshold)], fill=fill),
        )

    def _column_by_header(self, ws: Any, header: str) -> Optional[int]:
        header_row = self._header_row_index(ws)
        if header_row is None:
            return None
        for idx, cell in enumerate(ws[header_row], start=1):
            if cell.value == header:
                return idx
        return None

    def _header_row_index(self, ws: Any) -> Optional[int]:
        for row_idx in range(1, min(ws.max_row, 5) + 1):
            values = [cell.value for cell in ws[row_idx]]
            if any(value in self.PIPELINE_HEADERS or value in ("Metric", "severity", "class_name") for value in values):
                return row_idx
        return 1 if ws.max_row else None

    def _sheet_headers(self, ws: Any) -> List[Any]:
        if ws.max_row < 1:
            return []
        return [cell.value for cell in ws[1]]

    def _unique_sheet_title(self, wb: Workbook, base: str) -> str:
        title = base[:31]
        if title not in wb.sheetnames:
            return title
        index = 1
        while True:
            suffix = f"_{index}"
            title = f"{base[:31 - len(suffix)]}{suffix}"
            if title not in wb.sheetnames:
                return title
            index += 1

_excel_exporter: Optional[ExcelExporter] = None
_excel_exporter_lock = threading.Lock()


def get_excel_exporter(filename: str = "chess_robot_experiment.xlsx", subscribe: bool = True) -> ExcelExporter:
    """Return the runtime Excel exporter without import-time workbook mutation."""
    global _excel_exporter
    with _excel_exporter_lock:
        if _excel_exporter is None or _excel_exporter.filename != filename:
            _excel_exporter = ExcelExporter(filename=filename, subscribe=subscribe)
        elif subscribe:
            _excel_exporter.start_subscription()
        return _excel_exporter
